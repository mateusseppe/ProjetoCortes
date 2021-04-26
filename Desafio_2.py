import pulp
import openpyxl
from openpyxl import Workbook
import pandas as pd
import numpy as np

if __name__ == '__main__':
    I = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    l = [0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5, 5.5, 6]
    demanda = np.random.randint(10, 20, 12)
    l = dict(zip(I, l))  # comprimento dos blocos
    d = dict(zip(I, demanda))
    J = [1, 2, 3, 4, 5, 6]  # vetor das barras
    L = [9.5, 10, 10.5, 11, 11.5, 12]
    E = [12, 12, 12, 12, 12, 12]
    max_e = np.max(E)  # extraindo o maior estoque
    L = dict(zip(J, L))  # dict de comprimento das barras
    E = dict(zip(J, E))  # dict de estoques
    K = []
    for i in range(max_e):
        K.append(i + 1)
    triple_array = []
    for j in J:
        for i in I: 
            for k in K:
                triple_array.append((j,i,k))
    j_k_array = []
    for j in J:
        for k in K:
            j_k_array.append((j,k))
    # início do modelo
    model = pulp.LpProblem(
        "Problema do corte das barras com comprimento variável", pulp.LpMinimize)
    x = pulp.LpVariable.dicts("Padrão de corte", ((
        j, i, k) for j in J for i in I for k in K), lowBound=0, cat='Integer')
    y = pulp.LpVariable.dicts("Binário das barras", ((j, k)
                              for j in J for k in K), cat='Binary')
    model += pulp.lpSum(L[j]*y[j, k]
                        for j in J for k in K)
    for j in J:
        for k in K:
            model += pulp.lpSum(x[j, i, k]*l[i] for i in I) <= y[j, k]*L[j]
    for i in I:
        model += pulp.lpSum(x[j, i, k] for j in J for k in K) == d[i]
    for j in J:
        model += pulp.lpSum(y[j, k] for k in K) <= E[j]
    model.solve()

    # extração de soluções
    resultados = []
    for t in triple_array:
        resultados.append(x[t].varValue)
    resultados_y = []
    for t in j_k_array:
        resultados_y.append(y[t].varValue)
    resultados_y.extend(resultados) 
    y_resultado = resultados_y[0:len(J)*len(K)].copy()
    y_matriz = np.mat(y_resultado)
    y_matriz = y_matriz.reshape(len(K),len(J))
    y_matriz.shape
    x_resultado = resultados_y[len(J)*len(K):len(J)*len(K)+len(J)*len(K)*len(I)]
    x_matriz = np.mat(x_resultado)
    x_matriz = x_matriz.reshape(len(J),len(K)*len(I))
    x_matriz.shape
    ####
    wb = Workbook()
    ws = {}
    for j in range(len(J)):
        ws[j+1] = wb.create_sheet(str(L[j+1]))
        ws[j+1].title = str(L[j+1])
    rows = []
    for row in x_matriz:
        rows.append(row)
    for i in range(len(rows)):
        rows[i] = np.mat(rows[i])
        rows[i] = rows[i].reshape(len(I),len(K))
        rows[i].shape
        rows[i] = rows[i].tolist()
    for j in range(len(J)):
        for row in rows[j]:
            ws[j+1].append(row)
    wb.save('document.xlsx')

    soma_blocos = sum(l[i]*d[i] for i in I)
    print("A perda mínima é de:", pulp.value(
        model.objective) - soma_blocos, "metros")


